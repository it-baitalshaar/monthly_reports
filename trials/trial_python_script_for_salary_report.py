import calendar
from openpyxl import load_workbook
from openpyxl.styles import Protection, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
import os

# ---------------- CONFIG ----------------
TEMPLATE_FILE = "template.xlsx"
OUTPUT_EXCEL = "Payroll_Output.xlsx"
PDF_DIR = "pdf_cards"

YEAR = 2024
MONTH = 1  # January

employees = [
    {
        "name": "Ahmed Ali",
        "id": "EMP001",
        "daily_data": {
            1: {"att": "P", "hrs": 8},
            2: {"att": "P", "hrs": 8},
            3: {"att": "SL", "hrs": 0},
        }
    },
    {
        "name": "Mohammed Hassan",
        "id": "EMP002",
        "daily_data": {
            1: {"att": "P", "hrs": 9},
            2: {"att": "P", "hrs": 8},
        }
    }
]

# ---------------------------------------


def generate_worker_sheet(wb, template, emp):
    sheet = wb.copy_worksheet(template)
    sheet.title = emp["name"][:31]

    # Header
    sheet["B2"] = emp["name"]
    sheet["E2"] = emp["id"]
    sheet["F4"] = calendar.month_name[MONTH]

    days_in_month = calendar.monthrange(YEAR, MONTH)[1]
    start_row = 7

    for day in range(1, days_in_month + 1):
        r = start_row + day - 1
        sheet[f"A{r}"] = day

        data = emp["daily_data"].get(day, {})
        sheet[f"B{r}"] = data.get("att", "")
        sheet[f"C{r}"] = data.get("hrs", "")

        # Row formulas
        sheet[f"G{r}"] = f'=IF(B{r}="SL","SL","")'
        sheet[f"I{r}"] = (
            f'=IFERROR(IF(SUM(C{r}:G{r})>24,'
            f'"إجمالي > 24 ساعة.",SUM(C{r}:G{r})),"")'
        )

        # Ensure cell exists before setting alignment
        if sheet[f"J{r}"].value is None:
            sheet[f"J{r}"] = ""
        sheet[f"J{r}"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[r].height = 32

    # Lock formulas - unlock editable cells, keep formulas locked
    for r in range(7, 7 + days_in_month):
        for c in ['B', 'C', 'D', 'E', 'F', 'H', 'J']:
            cell = sheet[f"{c}{r}"]
            cell.protection = Protection(locked=False)
        # Keep formula cells (G and I) locked
        for c in ['G', 'I']:
            cell = sheet[f"{c}{r}"]
            cell.protection = Protection(locked=True)

    sheet.protection.sheet = True
    sheet.protection.enable()

    return sheet


def export_sheet_to_pdf(wb_path, sheet_name, filename):
    """Export a sheet to PDF. Requires workbook path to reload with calculated values."""
    os.makedirs(PDF_DIR, exist_ok=True)
    pdf_path = os.path.join(PDF_DIR, filename)

    try:
        # Reload workbook with data_only=True to get calculated formula values
        wb_data = load_workbook(wb_path, data_only=True)
        if sheet_name not in wb_data.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found for PDF export")
            return
        sheet = wb_data[sheet_name]
    except Exception as e:
        print(f"Error loading workbook for PDF export: {e}")
        return

    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=A4,
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    data = []
    # Get the actual data from the sheet (formulas are now calculated values)
    for row in sheet.iter_rows(min_row=1, max_row=40, max_col=10, values_only=True):
        row_data = [str(cell) if cell is not None else "" for cell in row]
        data.append(row_data)

    if not data:
        print(f"Warning: No data found in sheet {sheet_name}")
        return

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),  # Header row
    ]))

    try:
        doc.build([table])
    except Exception as e:
        print(f"Error building PDF: {e}")
        raise


def main():
    # Check if template file exists
    if not os.path.exists(TEMPLATE_FILE):
        print(f"❌ Error: Template file '{TEMPLATE_FILE}' not found!")
        return

    try:
        wb = load_workbook(TEMPLATE_FILE)
        
        # Check if Template sheet exists
        if "Template" not in wb.sheetnames:
            print(f"❌ Error: 'Template' sheet not found in '{TEMPLATE_FILE}'!")
            print(f"Available sheets: {', '.join(wb.sheetnames)}")
            return
        
        template = wb["Template"]
        generated_sheets = []

        # Generate sheets for each employee
        sheet_names = []
        for emp in employees:
            try:
                sheet = generate_worker_sheet(wb, template, emp)
                generated_sheets.append(sheet)
                sheet_names.append(sheet.title)
            except Exception as e:
                print(f"❌ Error generating sheet for {emp['name']}: {e}")
                continue

        # Save the workbook first
        wb.save(OUTPUT_EXCEL)

        # Export PDFs after saving (reload with calculated values)
        pdf_count = 0
        for sheet_name in sheet_names:
            try:
                pdf_name = f"{sheet_name}_{MONTH}_{YEAR}.pdf"
                export_sheet_to_pdf(OUTPUT_EXCEL, sheet_name, pdf_name)
                pdf_count += 1
            except Exception as e:
                print(f"❌ Error exporting PDF for {sheet_name}: {e}")
                continue

        print(f"✅ Excel generated: {OUTPUT_EXCEL}")
        print(f"✅ {pdf_count} PDF(s) exported to '{PDF_DIR}' directory")

    except FileNotFoundError:
        print(f"❌ Error: Template file '{TEMPLATE_FILE}' not found!")
    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    main()
