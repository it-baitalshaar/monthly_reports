import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Protection, Alignment
from datetime import datetime, timedelta
import calendar
import os
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm

# Database connection parameters
host = 'aws-0-ap-south-1.pooler.supabase.com'
port = 6543
dbname = 'postgres'
user = 'postgres.fhsvgeacwnnvqidyhnok'
password = 'Bait-Alshaar20'

def calculate_monthly_hours(from_date, to_date):
    """
    Calculate monthly hours based on number of days in the date range.
    Formula: number_of_days * 8 hours per day
    """
    delta = to_date - from_date
    number_of_days = delta.days + 1  # +1 to include both start and end dates
    monthly_hours = number_of_days * 8
    return monthly_hours, number_of_days

def get_date_input(prompt):
    """Get date input from user in YYYY-MM-DD format"""
    while True:
        date_str = input(prompt + " (YYYY-MM-DD): ").strip()
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            return date_obj
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD format (e.g., 2026-01-01)")

def get_monthly_hours_input(from_date, to_date):
    """Get monthly hours - either auto-calculated or manual input"""
    calculated_hours, num_days = calculate_monthly_hours(from_date, to_date)
    
    print(f"\nDate range: {from_date} to {to_date}")
    print(f"Number of days: {num_days}")
    print(f"Auto-calculated monthly hours: {calculated_hours} hours ({num_days} days × 8 hours/day)")
    
    while True:
        choice = input("\nUse auto-calculated hours? (y/n, default=y): ").strip().lower()
        if choice == '' or choice == 'y':
            return calculated_hours
        elif choice == 'n':
            try:
                manual_hours = float(input(f"Enter monthly hours (calculated: {calculated_hours}): ").strip())
                return manual_hours
            except ValueError:
                print("Invalid input. Please enter a number.")
        else:
            print("Please enter 'y' or 'n'")

def export_sheet_to_pdf(sheet, filename, output_dir):
    """
    Export an Excel sheet to PDF using reportlab
    Takes sheet object directly (matching payroll script pattern)
    """
    try:
        os.makedirs(output_dir, exist_ok=True)
        pdf_path = os.path.join(output_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        # Extract data from sheet
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=40, max_col=10):
            data.append([cell.value if cell.value else "" for cell in row])
        
        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        
        doc.build([table])
        return True
    except Exception as e:
        print(f"  Error exporting PDF for {sheet.title}: {e}")
        return False

def main():
    print("=" * 60)
    print("Monthly Attendance Report Generator")
    print("=" * 60)
    
    # Get date inputs
    print("\n--- Date Range Input ---")
    from_date = get_date_input("Enter FROM date")
    to_date = get_date_input("Enter TO date")
    
    if to_date < from_date:
        print("Error: TO date must be after FROM date!")
        return
    
    # Get monthly hours (auto or manual)
    monthly_hours = get_monthly_hours_input(from_date, to_date)
    
    # Template and output file paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, 'monthly_attendence_report_tempelate.xlsx')
    
    # Generate output filename based on dates
    output_filename = f"monthly_attendance_report_{from_date.strftime('%Y_%m_%d')}_to_{to_date.strftime('%Y_%m_%d')}.xlsx"
    output_path = os.path.join(script_dir, output_filename)
    
    # Check if template exists
    if not os.path.exists(template_path):
        print(f"\nError: Template file not found at: {template_path}")
        return
    
    print(f"\n--- Processing ---")
    print(f"Template: {template_path}")
    print(f"Output file: {output_path}")
    print(f"Monthly hours: {monthly_hours}")
    
    # Load template
    try:
        workbook = load_workbook(template_path)
        template_sheet = workbook['Sheet1']
        print("Template loaded successfully!")
    except Exception as e:
        print(f"Error loading template: {e}")
        return
    
    # Connect to the database
    connection = None
    try:
        connection = psycopg2.connect(
            host=host,
            port=port,
            dbname=dbname,
            user=user,
            password=password
        )
        cursor = connection.cursor()
        print("Connected to the database!")
        
        # Query with date range
        query = '''
            SELECT 
                e.employee_id,
                e.name,
                e.salary,
                a.date,
                a.notes,
                a.status,
                a.status_attendance,
                p.project_name,
                ap.working_hours,
                ap.overtime_hours
            FROM 
                "Employee" e
            LEFT JOIN 
                "Attendance" a ON e.employee_id = a.employee_id
            LEFT JOIN
                "Attendance_projects" ap ON a.id = ap.attendance_id
            LEFT JOIN
                projects p ON ap.project_id = p.project_id
            WHERE 
                a.date BETWEEN %s AND %s
            ORDER BY 
                e.employee_id, a.date;
        '''
        
        cursor.execute(query, (from_date, to_date))
        data = cursor.fetchall()
        
        if not data:
            print(f"\nNo attendance data found for the date range {from_date} to {to_date}")
            return
        
        df = pd.DataFrame(data, columns=[
            'employee_id', 'name', 'salary', 'date', 'notes', 'status',
            'status_attendance', 'project_name', 'working_hours', 'overtime_hours'
        ])
        
        grouped_data = df.groupby('employee_id')
        
        print(f"\nFound {len(grouped_data)} employees with attendance data")
        
        generated_sheets = []  # Track sheets for PDF export
        
        for employee_id, employee_data in grouped_data:
            employee_sheet = workbook.copy_worksheet(template_sheet)
            
            first_row = employee_data.iloc[0]
            employee_sheet['B4'] = first_row['name']
            employee_sheet.title = f"{first_row['name']}"
            employee_sheet.sheet_view.rightToLeft = template_sheet.sheet_view.rightToLeft
            employee_sheet['I2'] = first_row['salary']
            
            # Write date range and month name to the correct template cells
            employee_sheet['F2'] = from_date
            employee_sheet['F3'] = to_date
            employee_sheet['F4'] = calendar.month_name[from_date.month]
            
            # Apply cell protection (unlock editable cells)
            try:
                for r in range(7, 50):  # Adjust range based on your template
                    for c in ['B', 'C', 'D', 'E', 'F', 'H', 'J']:
                        try:
                            cell = employee_sheet[f"{c}{r}"]
                            cell.protection = Protection(locked=False)
                        except:
                            pass
            except:
                pass  # Skip if protection fails
            
            row_date = 7  # Start from row 7
            
            # Group by date so each day = exactly one row
            daily_grouped = employee_data.groupby('date')
            
            for date_value, day_data in daily_grouped:
                # These values are the same for all rows of the same day
                status = day_data.iloc[0]['status']
                status_attendance = day_data.iloc[0]['status_attendance']
                notes = day_data.iloc[0]['notes']
                
                # Write the date
                employee_sheet[f'A{row_date}'] = date_value
                
                # Write notes
                employee_sheet[f'J{row_date}'] = notes
                
                # Write attendance status code
                if status == 'present' and status_attendance == 'Present':
                    employee_sheet[f'B{row_date}'] = 'P'
                elif status == 'present' and status_attendance == 'Weekend':
                    employee_sheet[f'B{row_date}'] = 'W'
                elif status == 'present' and status_attendance == 'Holiday-Work':
                    employee_sheet[f'B{row_date}'] = 'H'
                elif status == 'absent' and status_attendance == 'Absence without excuse':
                    employee_sheet[f'B{row_date}'] = 'AWO'
                elif status == 'absent' and status_attendance == 'Sick Leave':
                    employee_sheet[f'B{row_date}'] = 'SL'
                elif status == 'absent' and status_attendance == 'Absence with excuse':
                    employee_sheet[f'B{row_date}'] = 'A'
                elif status == 'vacation':
                    employee_sheet[f'B{row_date}'] = 'V'
                
                # Sum total working hours for the day
                total_working_hours = day_data['working_hours'].sum()
                
                # If working hours are all 0 but it's Weekend or Holiday-Work, default to 8
                if total_working_hours == 0:
                    if status_attendance in ('Weekend', 'Holiday-Work'):
                        total_working_hours = 8
                
                if total_working_hours != 0:
                    employee_sheet[f'C{row_date}'] = total_working_hours
                
                # Sum overtime hours per type (Present / Weekend / Holiday-Work)
                total_overtime = day_data['overtime_hours'].sum()
                
                if total_overtime != 0:
                    if status_attendance == 'Present':
                        employee_sheet[f'D{row_date}'] = total_overtime
                    elif status_attendance == 'Weekend':
                        employee_sheet[f'E{row_date}'] = total_overtime
                    elif status_attendance == 'Holiday-Work':
                        employee_sheet[f'F{row_date}'] = total_overtime
                
                # Build the combined project string
                if len(day_data) == 1:
                    # Single project — just write the project name
                    project_text = day_data.iloc[0]['project_name']
                else:
                    # Multiple projects — combine with hours
                    project_parts = []
                    for _, proj_row in day_data.iterrows():
                        proj_name = proj_row['project_name'] if proj_row['project_name'] else 'Unknown'
                        proj_hours = int(proj_row['working_hours']) if proj_row['working_hours'] != 0 else 0
                        project_parts.append(f"{proj_name} {proj_hours}hrs")
                    project_text = " + ".join(project_parts)
                
                employee_sheet[f'H{row_date}'] = project_text
                
                # Set alignment for notes column
                try:
                    employee_sheet[f'J{row_date}'].alignment = Alignment(wrap_text=True)
                    employee_sheet.row_dimensions[row_date].height = 32
                except:
                    pass
                
                # Move to next row
                row_date += 1
            
            generated_sheets.append(employee_sheet)
            print(f"✓ Data written for Employee ID: {employee_id} — {first_row['name']}")
        
        # Save the workbook
        workbook.save(output_path)
        print(f"\n{'=' * 60}")
        print(f"✓ Excel file generated successfully!")
        print(f"✓ Output saved to: {output_path}")
        print(f"{'=' * 60}")
        
        # Ask if user wants to export PDFs
        export_pdf = input("\nExport PDFs for each employee? (y/n, default=n): ").strip().lower()
        if export_pdf == 'y':
            pdf_dir = os.path.join(script_dir, 'pdf_reports')
            print(f"\n--- Exporting PDFs ---")
            pdf_count = 0
            for sheet in generated_sheets:
                pdf_filename = f"{sheet.title}_{from_date.strftime('%Y_%m')}.pdf"
                if export_sheet_to_pdf(sheet, pdf_filename, pdf_dir):
                    pdf_count += 1
                    print(f"  ✓ {pdf_filename}")
            print(f"\n✓ {pdf_count} PDF(s) exported to: {pdf_dir}")
        
    except Exception as error:
        print(f"\nError: {error}")
        import traceback
        traceback.print_exc()
    
    finally:
        if connection:
            cursor.close()
            connection.close()
            print("\nDatabase connection closed.")

if __name__ == "__main__":
    main()
